# hello.py

from telebot import types
from zadachi_analit import test_analit_start
from zadachi_razrab import test_razrab_start
from openpyxl import load_workbook

def check_user_exists(user_id: int, file_name: str) -> bool:
    try:
        wb = load_workbook(filename=file_name)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            if row[0] == user_id:
                return True
        return False
    except Exception as e:
        print(f"Ошибка при проверке наличия пользователя в файле Excel: {e}")
        return False
# Функция для приветствия и получения имени пользователя
def handle_start(bot, message, user_data):
    user_id = message.from_user.id
    user_data[user_id] = {}
    user_data[user_id]['name'] = ''
    user_data[user_id]['surname'] = ''
    user_data[user_id]['age'] = ''
    user_data[user_id]['source'] = ''
    user_data[user_id]['nickname'] = ''
    if check_user_exists(user_id, 'analit_stag_list.xlsx') or check_user_exists(user_id, 'razrab_stag_list.xlsx'):
        bot.send_message(message.chat.id, "Вы уже проходили тест. Пожалуйста, дождитесь результатов.")
    else:
        user_data[user_id] = {}
        user_data[user_id]['name'] = ''
        user_data[user_id]['surname'] = ''
        user_data[user_id]['age'] = ''
        user_data[user_id]['source'] = ''
        user_data[user_id]['nickname'] = ''
        bot.send_photo(message.chat.id, open(r'cat.jpg', 'rb'))
        msg = bot.send_message(message.chat.id, "Приветствую! Я HR-бот компании Автомакон. Рад приветствовать тебя! 🌟 Для начала давай познакомимся. Начнем с имени?")
        bot.register_next_step_handler(msg, lambda m: get_name(bot, m, user_data))  # Не забыть заменить на get_name show_direction_description

def get_name(bot, message, user_data):
    user_id = message.from_user.id
    user_data[user_id]['name'] = message.text
    user_data[user_id]['nickname'] = message.from_user.username

    msg = bot.send_message(message.chat.id, f"Отлично, {user_data[user_id]['name']}! А какая у тебя фамилия?")
    bot.register_next_step_handler(msg, lambda m: get_surname(bot, m, user_data))

def get_surname(bot, message, user_data):
    user_id = message.from_user.id
    user_data[user_id]['surname'] = message.text

    msg = bot.send_message(message.chat.id, "Сколько тебе лет?")
    bot.register_next_step_handler(msg, lambda m: get_age(bot, m, user_data))

def get_age(bot, message, user_data):
    user_id = message.from_user.id
    user_data[user_id]['age'] = message.text

    msg = bot.send_message(message.chat.id, "Брр.. Так много! Мне то всего несколько месяцев! Откуда ты узнал(а) о наших стажировках?", reply_markup=create_source_keyboard())
    bot.register_next_step_handler(msg, lambda m: save_source(bot, m, user_data))

def create_source_keyboard():
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    keyboard.add(types.KeyboardButton("HeadHunter"), types.KeyboardButton("Сайт Автомакон"), types.KeyboardButton("Телеграм чаты"), types.KeyboardButton("Другое"))
    return keyboard

def save_source(bot, message, user_data):
    user_id = message.from_user.id
    user_data[user_id]['source'] = message.text

    msg = bot.send_message(message.chat.id, "Записываем для тебя ролик с нашим директором!")
    # Отправка видео
    video_url = r"test_video.mp4"
    bot.send_video(message.chat.id, open(video_url, 'rb'), reply_markup=create_video_keyboard())
    bot.register_next_step_handler(msg, lambda m: show_direction_description(bot, m, user_data))

def create_video_keyboard():
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    keyboard.add(types.KeyboardButton("Супер! Хочу к вам!"))
    return keyboard

def show_direction_description(bot, message, user_data):
    user_id = message.from_user.id
    # Описание направлений стажировки
    bot.send_message(message.chat.id, "Присоединяйся к нам на стажировку в Автомакон! У нас тебя ждут 4 месяца интенсивного обучения, классные наставники и возможность выбрать свой путь: Аналитика или Разработка. Дай карьере новый импульс с нами!")

    # Выбор направления с кнопками
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    keyboard.add(types.KeyboardButton("Разработка"), types.KeyboardButton("Аналитика"))
    msg = bot.send_message(message.chat.id, "Выбери направление стажировки и мы расскажем поподробнее:", reply_markup=keyboard)

    bot.register_next_step_handler(msg, lambda m: handle_direction_choice(bot, m, user_data))

def handle_direction_choice(bot, message, user_data):
    user_id = message.from_user.id
    direction = message.text

    if direction == "Разработка":
        bot.send_message(message.chat.id, "Разработка в Автомаконе - это возможность окунуться в мир программирования и создания инновационных технологических решений. Здесь ты освоишь современные языки программирования, научишься разрабатывать высококачественное программное обеспечение и применять передовые технологии в работе. Присоединяйся к нам и стань частью команды, создающей будущее!")
        ask_for_preference(bot, message, user_data, "разработка")
    elif direction == "Аналитика":
        bot.send_message(message.chat.id, "Системная аналитика в Автомаконе - это погружение в мир анализа бизнес-процессов и систем. Здесь ты научишься разбираться в сложных структурах, выявлять потребности бизнеса и предлагать эффективные технологические решения.")
        ask_for_preference(bot, message, user_data, "аналитика")

def ask_for_preference(bot, message, user_data, direction_name):
    user_id = message.from_user.id
    user_data[user_id]['direction'] = direction_name

    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    keyboard.add(types.KeyboardButton("Да, готов(а) к тесту"), types.KeyboardButton("Нет, вернуться к выбору"))
    msg = bot.send_message(message.chat.id, f"Нравится ли тебе направление {direction_name}? Ты готов проверить себя и узнать свой уровень знаний в этой области? Если да, то давай перейдем к тестированию по выбранному направлению! Ответь на вопросы и проверь свои знания. Помни, у тебя всего 30 минут!", reply_markup=keyboard)

    bot.register_next_step_handler(msg, lambda m: handle_preference_response(bot, m, user_data))

def handle_preference_response(bot, message, user_data):
    user_id = message.from_user.id
    response = message.text.lower()
    if response == "да, готов(а) к тесту":
        if user_data[user_id]['direction'] == "разработка":
            msg = bot.send_message(message.chat.id, "Точно? 30 минут! Вперед, Разработчики!")
            # Здесь вызываем функцию для начала теста для разработчиков
            bot.register_next_step_handler(msg, lambda m: test_razrab_start(bot, m, user_data))

        elif user_data[user_id]['direction'] == "аналитика":
            msg = bot.send_message(message.chat.id, "Точно? 30 минут! Вперед, Аналитики!")
            # Здесь вызываем функцию для начала теста для аналитиков
            bot.register_next_step_handler(msg, lambda m:test_analit_start(bot, m, user_data))
    elif response == "нет, вернуться к выбору":
        show_direction_description(bot, message, user_data)
    else:
        bot.send_message(message.chat.id, "Пожалуйста, ответьте 'Да' или 'Нет'.")
