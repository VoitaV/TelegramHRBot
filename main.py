# main.py

import telebot
import datetime
from config import token
from telebot.types import Message
from hello import handle_start
from zadachi_analit import test_analit_start, finish_test
from zadachi_razrab import test_razrab_start, finish_test
from openpyxl import load_workbook

bot_token = token
bot = telebot.TeleBot(bot_token)

# Словарь для хранения информации о пользователях
user_data = {}

@bot.message_handler(commands=['start'])
def start(message: Message):
    handle_start(bot, message, user_data)

# @bot.message_handler(func=lambda message: True)
# def handle_message(message: Message):
#     test_start(bot, message, user_data)

# # Обработчик завершения теста
# @bot.message_handler(func=finish_test)
# def handle_test_finished(message: Message):
#     save_to_excel(message, user_data)

def save_to_excel(message: Message, user_data: dict):
    user_id = message.from_user.id
    # Определяем файл Excel в зависимости от направления стажировки
    if user_data[user_id]['direction'].lower() == 'аналитика':
        file_name = 'analit_stag_list.xlsx'
    elif user_data[user_id]['direction'].lower() == 'разработка':
        file_name = 'razrab_stag_list.xlsx'
    else:
        return  # Если направление не определено, не сохраняем

    # Открываем или создаем файл Excel

    try:
        wb = load_workbook(filename=file_name)
        ws = wb.active
        # ws.append([
        #     "ID",
        #     "Ник в телеграм",
        #     "Имя",
        #     "Фамилия",
        #     "Возраст",
        #     "Откуда узнал",
        #     "Время начала теста",
        #     "Время конца теста",
        #     "Длительность теста",
        #     "Ответ на вопрос 1",
        #     "Ответ на вопрос 2",
        #     "Ответ на вопрос 3"
        # ])
        next_row = ws.max_row + 1
        # Добавляем данные в строку Excel
        row_data = [
            user_id,
            user_data[user_id]['nickname'],
            user_data[user_id]['name'],
            user_data[user_id]['surname'],
            user_data[user_id]['age'],
            user_data[user_id]['source'],
            datetime.datetime.fromtimestamp(user_data[user_id]['start_time']).strftime('%Y-%m-%d %H:%M:%S'),
            datetime.datetime.fromtimestamp(user_data[user_id]['end_time']).strftime('%Y-%m-%d %H:%M:%S'),
            user_data[user_id]['end_time'] - user_data[user_id]['start_time'],
        ]
        for i in range(3):
            answer = user_data[user_id][f'question{i + 1}_answer']
            row_data.append(answer)

        ws.append(row_data)

        # Сохраняем файл Excel
        wb.save(file_name)
        # bot.send_message(message.chat.id, f"Данные сохранены в файл {file_name}")
        if user_data[user_id]['score'] >= 3:
            bot.send_message(message.chat.id, f"Ты прекрасно справился! Мы скоро с тобой свяжемся!")
        elif user_data[user_id]['score'] >= 2:
            bot.send_message(message.chat.id, f"Ты можешь лучше! Мы скоро с тобой свяжемся!")
        else:
            bot.send_message(message.chat.id, f"Могло бы быть и получше, но не расстрайваися!")
                        
    except Exception as e:
        print(f"Ошибка при сохранении в файл Excel: {e}")

if __name__ == "__main__":
    bot.polling(none_stop=True)
